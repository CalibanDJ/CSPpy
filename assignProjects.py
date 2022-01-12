from pycsp3 import *
import openpyxl

dir_data = "data/"

# students first name
first_names = []

# students last name
names = []

# students specific school
schools = []

# students' interest for projects
wishes = []

# projects' name
projects = []


# parse data from excel file
def collect_data(file):
    wookbook = openpyxl.load_workbook(file)
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):
        if i == 0:
            for col in worksheet.iter_cols(4, worksheet.max_column):
                projects.append(col[i].value)
        else:
            cpt = 0
            wish = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if (cpt == 0):
                    names.append(col[i].value)
                elif (cpt == 1):
                    first_names.append(col[i].value)
                elif (cpt == 2):
                    schools.append(col[i].value)
                else:
                    wish.append(int(col[i].value))
                cpt += 1
            wishes.append(wish)


if not variant():
    # all the data - computation very long (38 students - 19 projects)
    collect_data(dir_data + "voeux-projets-semestriels-2021.xlsx")
elif variant("v1"):
    # 10 students - 5 projects
    collect_data(dir_data + "voeux-projets-semestriels-2021-light.xlsx")
elif variant("v2"):
    # 20 students - 10 projects
    collect_data(dir_data + "voeux-projets-semestriels-2021-light-2.xlsx")
elif variant("v3"):
    # 30 students - 15 projects
    collect_data(dir_data + "voeux-projets-semestriels-2021-light-3.xlsx")
elif variant("v4"):
    # 10 students - 8 projects
    collect_data(dir_data + "voeux-projets-semestriels-2021-light-more-project.xlsx")
elif variant("v5"):
    # 15 students - 5 projects
    collect_data(dir_data + "voeux-projets-semestriels-2021-light-more-students.xlsx")

nStudents = len(names)
nProjects = len(projects)
maxStudentsProject = 2

table = {(i, schools[i]) for i in range(nStudents)}

# project assigned to a student
p = VarArray(size=nStudents, dom=range(nProjects))

# cost of assigning a project to a student
c = VarArray(size=nStudents, dom=lambda i: wishes[i])

# preferences of ith student for the ith project
cp = VarArray(size=[nStudents, nProjects], dom=wishes)

s = VarArray(size=nStudents, dom={0, 1})
s2 = VarArray(size=nProjects, dom=range(-3, 3))
# [cp[i][j] == wishes[i][j] for j in range(nProjects) for i in range(nStudents)],
#   [c[i] == cp[i][p[i]] for i in range(nStudents)],
#   [Sum(p[j] == i for j in range(nStudents)) <= 2 for i in range(nProjects)]
x = Var(dom=range(50))
satisfy(
    # set
    [s[i] == (0 if schools[i] == 'EMMK' else 1) for i in range(nStudents)],
    #[Count(s, value=1)],
    #[Count(schools, value={'EMMK'}) == s[i] for i in range(nProjects)],
    #[s[i] == () for i in range(nProjects)],

    # assign preferences of ith student for the ith project
    [cp[i][j] == wishes[i][j] for j in range(nProjects) for i in range(nStudents)],

    # computing the cost of assigning the ith project to a student
    [cp[i][p[i]] == c[i] for i in range(nStudents)],

    # each projects is assigned at most 2 students
    Cardinality(p, occurrences={i: range(0, maxStudentsProject + 1) for i in range(nProjects)})
)

def equilibre(project):
    cpt = 0
    for i in range(nStudents):
        if p[i] == project :
            cpt += 1
    return cpt

minimize(
    # minimize the
    #Sum(c)
    Sum(equilibre(i) for i in range(nProjects))

)
